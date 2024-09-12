import time
from tendo import singleton
import psycopg2
import os
from dotenv import load_dotenv

load_dotenv()
import pandas as pd
import logging
from pymongo import MongoClient
import boto3
import requests
from sendgrid.helpers.mail import Mail
# from openpyxl.styles import Font, PatternFill,Alignment
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
import hashlib
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from datetime import datetime
import pytz
from openpyxl.styles import NamedStyle

ist = pytz.timezone('Asia/Kolkata')

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(lineno)d - %(message)s'
)

postgres_host = os.getenv("PG_HOST")
postgres_db = os.getenv("PG_DATABASE")
postgres_user = os.getenv("PG_USER")
postgres_password = os.getenv("PG_PASSWORD")
postgres_port = os.getenv("PG_PORT")

aws_access_key_id = os.getenv('AWS_ACCESS')
aws_secret_access_key = os.getenv('AWS_SECRET')
bucket_name = os.getenv('DEST_AWS_BUCKET_NAME')

SENDGRID_API_KEY = os.getenv('SENDGRID_API_KEY')

MONGO_URL = os.getenv('MONGO_URL')
client = MongoClient(MONGO_URL, maxIdleTimeMS=None)
logging.info("Mongo connection successful")

pgconn = psycopg2.connect(
    host=postgres_host,
    database=postgres_db,
    port=postgres_port,
    user=postgres_user,
    password=postgres_password
)

LIMIT = 1

currtime = int(time.time())
bucket_time = int(currtime / (90 * 24 * 60 * 60))


def try_convert_to_date(value):
    # List of common date and datetime formats to try
    date_formats = [
        "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%b-%Y", "%d %b %Y", "%Y%m%d",  # Date-only formats
        "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S",  # Date and time formats
        "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M", "%m/%d/%Y %H:%M",  # Without seconds
        "%d-%b-%Y %H:%M:%S", "%d %b %Y %H:%M:%S", "%d-%m-%Y"  # Other common variations
    ]
    for fmt in date_formats:
        try:
            return datetime.strptime(str(value), fmt), fmt  # Return the date and format
        except (ValueError, TypeError):
            continue
    return None, None  # Return None if conversion fails


def sendMailToClient(to_emails, template_id, dynamic_template_data):
    logging.info("sendMailToClient called...")
    api_key = SENDGRID_API_KEY
    url = 'https://api.sendgrid.com/v3/mail/send'

    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {api_key}',
    }

    for to_email in to_emails:
        message = Mail(
            from_email='info@finkraft.ai',
            to_emails=to_email,
        )
        message.template_id = template_id
        message.dynamic_template_data = dynamic_template_data

        try:
            # Convert the Mail object to JSON
            response = requests.post(
                url,
                headers=headers,
                json=message.get(),
                verify=False  # Disable SSL verification
            )
            logging.info(f"Email sent to {to_email} successfully! Status code: {response.status_code}")
        except Exception as e:
            logging.info(f"Error sending email to {to_email}: {e}")


def uploadFile(filename):
    try:
        logging.info("uploadFile called...")
        s3 = boto3.client('s3', aws_access_key_id=aws_access_key_id,
                          aws_secret_access_key=aws_secret_access_key)
        object = f"{bucket_time}/{filename}"
        s3_url = f"https://{bucket_name}.s3.amazonaws.com/{object}"
        s3.upload_file(filename,
                       bucket_name,
                       Key=object
                       )
        if os.path.exists(filename):
            os.remove(filename)
        return s3_url
    except Exception as e:
        logging.info("Exception happened in the getS3Url: " + str(e))
        if os.path.exists(filename):
            os.remove(filename)
        return None


def findFileHash(file_path, hash_algo='sha256'):
    hash_func = hashlib.new(hash_algo)
    with open(file_path, 'rb') as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_func.update(chunk)
    return hash_func.hexdigest()


def getCreatedBy(createdbyId):
    logging.info("getCreatedBy is called...")
    with pgconn.cursor() as cursor:
        select_query = f"SELECT name FROM users WHERE id='{createdbyId}'"
        logging.info("Query: " + str(select_query))
        cursor.execute(select_query)
        results = cursor.fetchall()
        if results is None or len(results) == 0:
            return createdbyId
        else:
            for row in results:
                return row[0]


def xlsxWriter(filename, schemaName, header_details, hyperlink_headers, date_headers, createdbyId, df):
    try:
        cover_image_path = './finkraftlogo.png'
        # Create the cover sheet
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            wb = writer.book
            cover_ws = wb.create_sheet(title='Cover')

            # Merge columns A to D and rows 1 to 4
            cover_ws.merge_cells('A1:D4')

            # Add image to cover sheet
            if cover_image_path and os.path.isfile(cover_image_path):
                img = Image(cover_image_path)

                # Calculate width of columns A to D
                total_width = (cover_ws.column_dimensions['A'].width or 10) + \
                              (cover_ws.column_dimensions['B'].width or 10) + \
                              (cover_ws.column_dimensions['C'].width or 10) + \
                              (cover_ws.column_dimensions['D'].width or 10)

                # Calculate height of rows 1 to 4
                total_height = (cover_ws.row_dimensions[1].height or 15) + \
                               (cover_ws.row_dimensions[2].height or 15) + \
                               (cover_ws.row_dimensions[3].height or 15) + \
                               (cover_ws.row_dimensions[4].height or 15)

                # Adjust the image size to fit the merged cells
                img.width = total_width * 7.5  # Approximate width in pixels
                img.height = total_height * 1.2  # Approximate height in pixels

                # Place the image in the merged area (A1)
                cover_ws.add_image(img, 'A1')

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Apply the border to the merged cells (A1:D4)
            for row in cover_ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=4):
                for cell in row:
                    cell.border = thin_border

            cover_ws.merge_cells('A7:D7')
            schemaNameCellStart = 7
            schemaNameCell = cover_ws.cell(row=schemaNameCellStart, column=1, value=schemaName)
            schemaNameCell.font = Font(size=16, bold=True)
            schemaNameCell.alignment = Alignment(horizontal='center', vertical='center')

            createdInfoCellStart = 8
            createdAtCell = cover_ws.cell(row=createdInfoCellStart, column=1, value="Created On")
            createdAtCell.font = Font(size=12, bold=True)

            current_time_ist = datetime.now(ist)
            current_date_formatted = current_time_ist.strftime("%d/%m/%Y")
            createdAtValCell = cover_ws.cell(row=createdInfoCellStart, column=2, value=current_date_formatted)
            createdAtValCell.font = Font(size=12, bold=True)

            createdByCell = cover_ws.cell(row=createdInfoCellStart, column=3, value="Created By")
            createdByCell.font = Font(size=12, bold=True)

            createdBy = getCreatedBy(createdbyId)
            createdByValCell = cover_ws.cell(row=createdInfoCellStart, column=4, value=createdBy)
            createdByValCell.font = Font(size=12, bold=True)

            mainheader_color = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            for row in cover_ws.iter_rows(min_row=schemaNameCellStart, max_row=createdInfoCellStart, min_col=1,
                                          max_col=4):
                for cell in row:
                    cell.border = thin_border
                    cell.fill = mainheader_color

            headerCellStart = 10
            header_row = ["Column Name", "Data Type", "Agg Type", "Agg Value", "Group Name", "UI Format Type",
                          "Grouping", "Pivot"]
            for col_idx, value in enumerate(header_row, start=1):
                cell = cover_ws.cell(row=headerCellStart, column=col_idx, value=value)
                cell.font = Font(size=13, bold=True)
            header_color = PatternFill(start_color="6AA84F", end_color="6AA84F", fill_type="solid")
            for row in cover_ws.iter_rows(min_row=headerCellStart, max_row=headerCellStart, min_col=1, max_col=8):
                for cell in row:
                    cell.border = thin_border
                    cell.fill = header_color

            dataCellStart = 11
            dataCellEnd = 11
            for row_idx, row_data in enumerate(header_details, start=dataCellStart):
                cell_a = cover_ws.cell(row=row_idx, column=1, value=row_data[0])
                cell_a.font = Font(bold=True, size=10)
                cover_ws.cell(row=row_idx, column=2, value=row_data[1])
                cover_ws.cell(row=row_idx, column=3, value=row_data[2])
                cover_ws.cell(row=row_idx, column=4, value=row_data[3])
                cover_ws.cell(row=row_idx, column=5, value=row_data[4])
                cover_ws.cell(row=row_idx, column=6, value=row_data[5])
                cover_ws.cell(row=row_idx, column=7, value=row_data[6])
                cover_ws.cell(row=row_idx, column=8, value=row_data[7])
                dataCellEnd += 1

            header_details.append(header_row)
            for col_idx in range(1, 8):  # Columns A, B, C (1-based index)
                max_length = 0
                for row_idx, row_data in enumerate(header_details, start=dataCellStart):
                    if len(str(row_data[col_idx - 1])) > max_length:
                        max_length = len(str(row_data[col_idx - 1]))

                # Set column width (slightly larger to accommodate characters)
                adjusted_width = max_length + 4  # Add some padding for better spacing
                cover_ws.column_dimensions[
                    chr(64 + col_idx)].width = adjusted_width  # chr(64 + col_idx) converts to 'A', 'B', 'C'

            data_color = PatternFill(start_color="FFEBD6", end_color="FFEBD6", fill_type="solid")
            for row in cover_ws.iter_rows(min_row=dataCellStart, max_row=dataCellEnd - 1, min_col=1, max_col=8):
                for cell in row:
                    cell.border = thin_border
                    cell.fill = data_color

            for col in df.columns:
                if isinstance(df[col], pd.Series):
                    try:
                        if pd.api.types.is_numeric_dtype(df[col]):
                            continue  # Already numeric, no need to convert
                        df[col] = pd.to_numeric(df[col])
                    except ValueError:
                        logging.warning(f"Skipping non-numeric column: {col}")
                        continue
                else:
                    logging.warning(f"Skipping column {col} as it is not a valid Series")

            date_style = NamedStyle(name="date", number_format='DD/MM/YYYY')
            datetime_style = NamedStyle(name="datetime", number_format='DD/MM/YYYY HH:MM:SS')

            # Split data into chunks of 10,000 rows
            for chunk_num, chunk in enumerate(range(0, len(df), 10000)):
                sheet_name = f"Sheet{chunk_num + 1}"
                df_chunk = df.iloc[chunk:chunk + 10000]
                df_chunk.to_excel(writer, sheet_name=sheet_name, index=False)

                # Open the workbook to format the header
                ws = wb[sheet_name]

                # Set header style (font size 12, yellow fill)
                for cell in ws[1]:
                    cell.font = Font(size=12, bold=True)
                    cell.fill = PatternFill(start_color="f5f6f9", end_color="f5f6f9", fill_type="solid")
                ws.auto_filter.ref = ws.dimensions

                # Identify columns that should be hyperlinks based on header names
                link_header_row = [cell.value for cell in ws[1]]
                hyperlink_columns = [link_header_row.index(header) + 1 for header in hyperlink_headers if
                                     header in link_header_row]

                for col_idx in hyperlink_columns:
                    for row in range(2, len(df_chunk) + 2):  # Skip header row, starting from row 2
                        cell_value = df_chunk.iloc[row - 2, col_idx - 1]
                        if pd.notna(cell_value):
                            cell = ws.cell(row=row, column=col_idx, value=cell_value)  # Set the display text
                            cell.hyperlink = cell_value  # Set the URL as the hyperlink
                            cell.style = "Hyperlink"

                date_header_row = [cell.value for cell in ws[1]]
                date_columns = [date_header_row.index(header) + 1 for header in date_headers if
                                header in date_header_row]

                for col_idx in date_columns:
                    for row in range(2, len(df_chunk) + 2):  # Skip header row, starting from row 2
                        cell_value = df_chunk.iloc[row - 2, col_idx - 1]
                        if pd.notna(cell_value):
                            date_value, format_used = try_convert_to_date(cell_value)
                            cell = ws.cell(row=row, column=col_idx,
                                           value=cell_value)  # Keep original value if not converted
                            if date_value:
                                cell.value = date_value
                                # Apply datetime format if time is included, otherwise apply date format
                                if "%H:%M" in format_used or "%H:%M:%S" in format_used:
                                    cell.style = datetime_style
                                else:
                                    cell.style = date_style

        return True
    except Exception as e:
        logging.info("Exception happened in the xlsxWriter: " + str(e))
        return False


def getData(tablename, wsname, columnDefs, columnMapping, schemaName, createdbyId):
    try:
        logging.info("getData called...")
        filename = '_'.join(wsname) + "_" + str(currtime) + ".xlsx"
        total_records = 0
        column_query = ""
        header_details = []
        columnMap = {}
        hyperlink_headers = []
        date_columns = []
        header_details_map = {}

        for i in range(len(columnMapping)):
            aggtype = None
            if columnMapping[i]["data_type"] == "NUMBER":
                aggtype = "SUM"
            else:
                aggtype = "UNIQUE COUNT"
            columnMap[columnMapping[i]["destination_key"]] = [columnMapping[i]["data_type"], aggtype]

        for i in range(len(columnDefs)):
            columnadded = False
            if "hide" in columnDefs[i] and columnDefs[i]["hide"] == False:
                column_query = column_query + '"' + columnDefs[i]["field"] + '" AS "' + columnDefs[i][
                    "headerName"] + '"'
                column_details = columnMap.get(columnDefs[i]["field"])
                columnadded = True
                uiFormatType = ""
                if "formatType" in columnDefs[i] and columnDefs[i]["formatType"] != "NONE":
                    uiFormatType = columnDefs[i]["formatType"]

                rowGrouping = "NO"
                if "enableRowGroup" in columnDefs[i] and columnDefs[i]["enableRowGroup"] == True:
                    rowGrouping = "YES"

                enablePivot = "NO"
                if "enablePivot" in columnDefs[i] and columnDefs[i]["enablePivot"] == True:
                    enablePivot = "YES"

                header_details_map[columnDefs[i]["headerName"]] = [columnDefs[i]["headerName"], column_details[0],
                                                                   column_details[1], "", uiFormatType, rowGrouping,
                                                                   enablePivot]

                if "formatType" in columnDefs[i] and columnDefs[i]["formatType"] == "INVOICE":
                    hyperlink_headers.append(columnDefs[i]["headerName"])

                if ("formatType" in columnDefs[i] and columnDefs[i]["formatType"] == "DATE") or (
                        "filter" in columnDefs[i] and columnDefs[i]["filter"] == "agDateColumnFilter"):
                    date_columns.append(columnDefs[i]["headerName"])

                # if i < len(columnDefs) - 1:
                #     column_query += ", "
            elif "children" in columnDefs[i]:
                for j in range(len(columnDefs[i]["children"])):
                    columnadded = False
                    if "hide" in columnDefs[i]["children"][j] and columnDefs[i]["children"][j]["hide"] == False:
                        column_query = column_query + '"' + columnDefs[i]["children"][j]["field"] + '" AS "' + \
                                       columnDefs[i]["children"][j]["headerName"] + '"'
                        column_details = columnMap.get(columnDefs[i]["children"][j]["field"])
                        columnadded = True
                        uiFormatType = ""
                        if "formatType" in columnDefs[i]["children"][j] and columnDefs[i]["children"][j][
                            "formatType"] != "NONE":
                            uiFormatType = columnDefs[i]["children"][j]["formatType"]

                        rowGrouping = "NO"
                        if "enableRowGroup" in columnDefs[i]["children"][j] and columnDefs[i]["children"][j][
                            "enableRowGroup"] == True:
                            rowGrouping = "YES"

                        enablePivot = "NO"
                        if "enablePivot" in columnDefs[i]["children"][j] and columnDefs[i]["children"][j][
                            "enablePivot"] == True:
                            enablePivot = "YES"

                        header_details_map[columnDefs[i]["children"][j]["headerName"]] = [
                            columnDefs[i]["children"][j]["headerName"], column_details[0],
                            column_details[1], columnDefs[i]["headerName"], uiFormatType, rowGrouping,
                            enablePivot]

                        if "formatType" in columnDefs[i]["children"][j] and columnDefs[i]["children"][j][
                            "formatType"] == "INVOICE":
                            hyperlink_headers.append(columnDefs[i]["children"][j]["headerName"])
                        if ("formatType" in columnDefs[i]["children"][j] and columnDefs[i]["children"][j][
                            "formatType"] == "DATE") or (
                                "formatType" in columnDefs[i]["children"][j] and columnDefs[i]["children"][j][
                            "filter"] == "agDateColumnFilter"):
                            date_columns.append(columnDefs[i]["children"][j]["headerName"])

                        if j < len(columnDefs[i]["children"]) - 1 and columnadded == True:
                            column_query += ", "
            if i < len(columnDefs) - 1 and columnadded == True:
                column_query += ", "

        df = None
        for i in range(len(wsname)):
            with pgconn.cursor() as cursor:
                select_query = f"SELECT {column_query} FROM {tablename} WHERE \"Workspace\" ILIKE %s"
                logging.info("Query: " + str(select_query))
                logging.info("Param: " + str(wsname[i]))
                cursor.execute(select_query, (wsname[i],))
                results = cursor.fetchall()
                if results is None or len(results) == 0:
                    continue
                logging.info("No. of record: " + str(len(results)))
                column_names = [desc[0] for desc in cursor.description]
            df_temp = pd.DataFrame(results, columns=column_names)
            df = pd.concat([df, df_temp], ignore_index=True)

        header_details = []

        for key, val in header_details_map.items():
            count = 0
            if val[2] == "SUM":
                count = df[str(key)].sum()
            elif val[2] == "UNIQUE COUNT":
                count = df[str(key)].nunique()
            header_details.append([key, val[1], val[2], count, val[3], val[4], val[5], val[6]])

        if df is not None and len(df) != 0:
            total_records += len(df)
            status = xlsxWriter(filename, schemaName, header_details, hyperlink_headers, date_columns, createdbyId, df)
            if not status:
                return None, None, None
        filehash = findFileHash(filename)
        logging.info(f"Total records written: {total_records}")
        return filename, total_records, filehash
    except Exception as e:
        logging.info("Exception happened in the getData: " + str(e))
        return None, None, None


def getWorkspaceName(workspaceids):
    logging.info("getWorkspaceName called...")
    finalresult = []
    workspaceids = tuple(workspaceids)
    with pgconn.cursor() as cursor:
        select_query = "SELECT name FROM workspaces WHERE id in %s"
        cursor.execute(select_query, (workspaceids,))
        results = cursor.fetchall()
        for row in results:
            finalresult.append(row[0])
        return finalresult


def removeFile(filename):
    logging.info("removeFile called...")
    if os.path.exists(filename):
        os.remove(filename)


def getPendingJob():
    logging.info("getPendingJob called...")
    db = client['gstservice']
    collection = db['recon_report']
    result = list(collection.find({"reportId": "414687f1-1aa5-4cdd-9166-8ebae1cfad7d"}).limit(LIMIT))
    # result = list(collection.find({"status": "PENDING"}).sort({"createdBy": -1}).limit(LIMIT))
    return result


def getSchema(moduleId):
    logging.info("getSchema called...")
    db = client['gstservice']
    collection = db['AG_TABLE_SCHEMA']
    result = collection.find_one({"moduleId": moduleId})
    return result


if __name__ == '__main__':
    try:
        me = singleton.SingleInstance()
        logging.info("======================================================")
        jobs = getPendingJob()

        db = client['gstservice']
        collection = db['recon_report']

        if jobs is not None and len(jobs) != 0:
            for i in range(len(jobs)):
                logging.info("Processing for job: " + str(jobs[i]))
                if "workspace_id" in jobs[i] and "table_name" in jobs[i]:
                    workspacename = getWorkspaceName(jobs[i]['workspace_id'])
                    logging.info("Workspace Names: " + str(workspacename))
                    if workspacename is not None and len(workspacename) != 0:
                        schemaDetails = getSchema(jobs[i]["moduleId"])
                        logging.info("Schema Details: " + str(schemaDetails))
                        if schemaDetails is None:
                            key_to_check = {"_id": jobs[i]["_id"]}
                            result = collection.update_one(
                                key_to_check,
                                {
                                    "$set": {
                                        "status": "SCHEMA DETAILS MISSING",
                                        "total_record": 0
                                    }
                                })
                            if result.matched_count > 0:
                                logging.info("Updated the document: " + str(key_to_check))
                            else:
                                logging.info("No updates for the document: " + str(key_to_check))
                            continue

                        filename, count, filehash = getData(jobs[i]['table_name'], workspacename,
                                                            schemaDetails["state"]["columnDefs"],
                                                            schemaDetails["state"]["columnMapping"],
                                                            schemaDetails["name"], jobs[i]["createdBy"])
                        logging.info("Filename: " + str(filename))
                        logging.info("Total no. of records: " + str(count))
                        if count is None:
                            key_to_check = {"_id": jobs[i]["_id"]}
                            result = collection.update_one(
                                key_to_check,
                                {
                                    "$set": {
                                        "status": "EXCEPTION IN GETDATA",
                                        "total_record": 0
                                    }
                                })
                            if result.matched_count > 0:
                                logging.info("Updated the document: " + str(key_to_check))
                            else:
                                logging.info("No updates for the document: " + str(key_to_check))

                        elif count != 0:
                            s3_url = uploadFile(filename)
                            logging.info("S3 URL: " + str(s3_url))
                            if s3_url is not None:
                                subject = ""
                                if "report_name" in jobs[i]:
                                    subject = subject + str(jobs[i]["report_name"]) + " is ready to download"
                                else:
                                    subject = "Report ready to download"

                                dynamic_template_data = {
                                    "subject": subject,
                                    "description": "As per your request we have generated this report of your workspace.",
                                    "download_link": "https://files.finkraft.ai/report-" + str(filehash),
                                }
                                template_id = "d-a6a5853662824aa7a69e990013cf1faa"
                                to_emails = []
                                if "to_emails" in jobs[i]:
                                    to_emails = jobs[i]["to_emails"]
                                if len(to_emails) != 0:
                                    sendMailToClient(to_emails, template_id, dynamic_template_data)
                                    key_to_check = {"_id": jobs[i]["_id"]}
                                    result = collection.update_one(
                                        key_to_check,
                                        {
                                            "$set": {
                                                "status": "COMPLETED",
                                                "link": s3_url,
                                                "total_record": count,
                                                "filehash": filehash
                                            }
                                        })
                                    if result.matched_count > 0:
                                        logging.info("Updated the document: " + str(key_to_check))
                                    else:
                                        logging.info("No updates for the document: " + str(key_to_check))
                                else:
                                    key_to_check = {"_id": jobs[i]["_id"]}
                                    result = collection.update_one(
                                        key_to_check,
                                        {
                                            "$set": {
                                                "status": "TO MAIL MISSING",
                                                "total_record": count
                                            }
                                        })
                                    if result.matched_count > 0:
                                        logging.info("Updated the document: " + str(key_to_check))
                                    else:
                                        logging.info("No updates for the document: " + str(key_to_check))
                            else:
                                key_to_check = {"_id": jobs[i]["_id"]}
                                result = collection.update_one(
                                    key_to_check,
                                    {
                                        "$set": {
                                            "status": "FAILED TO GENERATE LINK",
                                            "total_record": count
                                        }
                                    })
                                if result.matched_count > 0:
                                    logging.info("Updated the document: " + str(key_to_check))
                                else:
                                    logging.info("No updates for the document: " + str(key_to_check))
                        elif count == 0:
                            key_to_check = {"_id": jobs[i]["_id"]}
                            result = collection.update_one(
                                key_to_check,
                                {
                                    "$set": {
                                        "status": "COMPLETED",
                                        "total_record": count
                                    }
                                })
                            if result.matched_count > 0:
                                logging.info("Updated the document: " + str(key_to_check))
                            else:
                                logging.info("No updates for the document: " + str(key_to_check))

                        removeFile(filename)
                    else:
                        key_to_check = {"_id": jobs[i]["_id"]}
                        result = collection.update_one(
                            key_to_check,
                            {
                                "$set": {
                                    "status": "WORKSPACE NOT FOUND"
                                }
                            })
                        if result.matched_count > 0:
                            logging.info("Updated the document: " + str(key_to_check))
                        else:
                            logging.info("No updates for the document: " + str(key_to_check))
                else:
                    key_to_check = {"_id": jobs[i]["_id"]}
                    result = collection.update_one(
                        key_to_check,
                        {
                            "$set": {
                                "status": "WORKSPACE ID OR TABLE NAME MISSING"
                            }
                        })
                    if result.matched_count > 0:
                        logging.info("Updated the document: " + str(key_to_check))
                    else:
                        logging.info("No updates for the document: " + str(key_to_check))
        else:
            logging.info("No pending jobs")
        logging.info("======================================================")
    except Exception as e:
        logging.info("Exception happened in the main: " + str(e))
